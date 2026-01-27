"""
Page Streamlit pour la facturation
"""
import streamlit as st
import pandas as pd
from datetime import datetime
from pathlib import Path
from facturation.utils import (
    load_factures, save_facture, load_config,
    generate_numero_facture, is_facture_locked, get_excel_path,
    init_excel_if_needed
)
from facturation.pdf_generator import generate_invoice_pdf


def load_clients_data():
    """
    Charge les données clients depuis la base existante
    Retourne un DataFrame avec les colonnes nécessaires
    """
    try:
        from whatsapp_database.queries import get_artisans
        
        artisans = get_artisans(limit=10000)
        if artisans and len(artisans) > 0:
            df = pd.DataFrame(artisans)
            clients_df = pd.DataFrame({
                'nom': df.get('nom_entreprise', df.get('nom', '')),
                'adresse': df.get('adresse', ''),
                'email': df.get('email', ''),
                'telephone': df.get('telephone', ''),
                'siret': df.get('siret', ''),
                'ville': df.get('ville', ''),
                'code_postal': df.get('code_postal', '')
            })
            clients_df['adresse_complete'] = clients_df.apply(
                lambda row: f"{row.get('adresse', '')}, {row.get('code_postal', '')} {row.get('ville', '')}".strip(', '),
                axis=1
            )
            valid_clients = clients_df[clients_df['nom'].notna() & (clients_df['nom'] != '')]
            return valid_clients
    except Exception:
        pass
    
    return pd.DataFrame()


def render_facturation_page():
    """Affiche l'onglet de facturation"""
    
    st.title("📄 Facturation")
    
    # Initialiser le fichier Excel si nécessaire
    init_excel_if_needed()

    # Charger les données sans cache pour toujours avoir les données à jour
    factures_df = load_factures()
    config = load_config()
    clients_df = load_clients_data()
    
    # Debug amélioré
    if len(factures_df) == 0:
        excel_path = get_excel_path()
        invoices_dir = Path(__file__).parent.parent / "invoices"
        
        # Vérifier s'il y a des PDFs dans invoices/
        pdf_count = 0
        if invoices_dir.exists():
            pdf_count = len(list(invoices_dir.rglob("*.pdf")))
        
        if excel_path.exists():
            # Vérifier le contenu réel du fichier
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                if 'FACTURES' in wb.sheetnames:
                    sheet = wb['FACTURES']
                    row_count = sheet.max_row
                    st.warning(f"⚠️ Aucune facture trouvée dans l'Excel")
                    st.info(f"📊 L'onglet FACTURES contient {row_count} lignes (en-têtes inclus)")
                    
                    if pdf_count > 0:
                        st.warning(f"📄 Mais {pdf_count} PDF(s) trouvé(s) dans {invoices_dir}")
                        st.info("💡 Les PDFs existent mais les données ne sont pas dans l'Excel. Créez une nouvelle facture pour initialiser l'Excel correctement.")
                    
                    # Afficher un bouton pour réessayer
                    if st.button("🔄 Réessayer le chargement"):
                        st.experimental_rerun()
                else:
                    st.error(f"❌ L'onglet 'FACTURES' n'existe pas dans {excel_path}")
            except Exception as e:
                st.error(f"❌ Erreur lors de la vérification du fichier: {e}")
                st.info(f"📁 Le fichier existe: {excel_path}")
        else:
            if pdf_count > 0:
                st.warning(f"📄 {pdf_count} PDF(s) trouvé(s) dans {invoices_dir} mais le fichier Excel n'existe pas")
                st.info("💡 Créez une nouvelle facture pour initialiser l'Excel. Les anciens PDFs ne seront pas automatiquement importés.")
            else:
                st.info(f"📝 Le fichier Excel n'existe pas encore. Il sera créé lors de la première facture.")
    
    # Formulaire de création de facture
    st.header("➕ Créer une nouvelle facture")
    
    # Valeurs par défaut pour pré-remplissage
    default_values = {
        'client_nom': '',
        'client_adresse': '',
        'client_email': '',
        'client_siret': '',
        'client_ref': ''
    }
    
    # Sélection du client (en dehors du form)
    if not clients_df.empty:
        client_options = ['-- Nouveau client --'] + list(clients_df['nom'].unique())
        selected_client = st.selectbox(
            "Sélectionner un client",
            client_options,
            key="client_select"
        )
        
        if selected_client != '-- Nouveau client --':
            client_data = clients_df[clients_df['nom'] == selected_client].iloc[0]
            default_values['client_nom'] = client_data.get('nom', '')
            default_values['client_adresse'] = client_data.get('adresse_complete', client_data.get('adresse', ''))
            default_values['client_email'] = client_data.get('email', '')
            # Ne pas afficher None pour le SIRET
            siret_value = client_data.get('siret', '')
            if siret_value is None or str(siret_value).strip().lower() in ['none', 'nan', 'null', '']:
                default_values['client_siret'] = ''
            else:
                default_values['client_siret'] = str(siret_value).strip()
            default_values['client_ref'] = selected_client
    
    # Formulaire
    with st.form("form_facture", clear_on_submit=True):
        col1, col2 = st.columns(2)
        
        with col1:
            st.text_input(
                "Nom / Raison sociale *",
                value=default_values['client_nom'],
                key="form_client_nom"
            )
            st.text_area(
                "Adresse *",
                value=default_values['client_adresse'],
                height=80,
                key="form_client_adresse"
            )
            st.text_input(
                "Email",
                value=default_values['client_email'],
                key="form_client_email"
            )
            st.text_input(
                "SIRET",
                value=default_values['client_siret'],
                key="form_client_siret"
            )
        
        with col2:
            st.text_area(
                "Description de la prestation *",
                height=80,
                key="form_description"
            )
            
            st.date_input(
                "Date d'émission",
                value=datetime.now().date(),
                key="form_date_emission"
            )
        
        # Quantité et Prix unitaire sur une seule ligne (en dehors des colonnes principales)
        col_qty, col_rate = st.columns(2)
        with col_qty:
            st.number_input(
                "Quantité *",
                min_value=1.0,
                step=1.0,
                value=1.0,
                format="%.0f",
                key="form_quantite"
            )
        
        with col_rate:
            st.number_input(
                "Prix unitaire HT (€) *",
                min_value=0.0,
                step=0.01,
                format="%.2f",
                key="form_prix_unitaire"
            )
        
        submit_button = st.form_submit_button("✅ Générer la facture")
    
    # Traitement après soumission
    if submit_button:
        # Récupérer les valeurs
        client_nom = st.session_state.get('form_client_nom', '')
        client_adresse = st.session_state.get('form_client_adresse', '')
        client_email = st.session_state.get('form_client_email', '')
        client_siret_raw = st.session_state.get('form_client_siret', '')
        # Nettoyer le SIRET pour éviter "None"
        if client_siret_raw is None or str(client_siret_raw).strip().lower() in ['none', 'nan', 'null', '']:
            client_siret = ''
        else:
            client_siret = str(client_siret_raw).strip()
        
        description = st.session_state.get('form_description', '')
        quantite = st.session_state.get('form_quantite', 1.0)
        prix_unitaire = st.session_state.get('form_prix_unitaire', 0.0)
        date_emission = st.session_state.get('form_date_emission', datetime.now().date())
        
        # Calculer le montant total
        montant_total = float(quantite) * float(prix_unitaire)
        
        # Validation - seulement les champs obligatoires
        errors = []
        if not client_nom or client_nom.strip() == '':
            errors.append("Nom / Raison sociale")
        if not client_adresse or client_adresse.strip() == '':
            errors.append("Adresse")
        if not description or description.strip() == '':
            errors.append("Description de la prestation")
        if quantite <= 0:
            errors.append("Quantité")
        if prix_unitaire <= 0:
            errors.append("Prix unitaire HT")
        
        if errors:
            st.error(f"⚠️ Veuillez remplir les champs obligatoires suivants : {', '.join(errors)}")
        else:
            try:
                numero = generate_numero_facture()
                
                facture_data = {
                    'numero': numero,
                    'date_emission': pd.Timestamp(date_emission),
                    'client_nom': client_nom,
                    'client_ref': default_values['client_ref'] if default_values['client_ref'] else client_nom,
                    'client_adresse': client_adresse,
                    'client_email': client_email,
                    'client_siret': client_siret,
                    'description': description,
                    'quantite': float(quantite),
                    'prix_unitaire': float(prix_unitaire),
                    'montant': montant_total,
                    'statut': 'brouillon',
                    'chemin_pdf': '',
                    'created_at': pd.Timestamp.now(),
                    'locked': False
                }
                
                with st.spinner("Génération du PDF..."):
                    chemin_pdf = generate_invoice_pdf(facture_data, config)
                    facture_data['chemin_pdf'] = str(chemin_pdf)
                
                save_facture(facture_data)
                
                st.success(f"✅ Facture {numero} créée avec succès !")
                st.info(f"📄 PDF généré : {chemin_pdf}")
                st.experimental_rerun()
                
            except Exception as e:
                st.error(f"❌ Erreur lors de la création : {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    
    st.markdown("---")

    # Tableau des factures existantes
    st.header("📊 Liste des factures")
    
    if not factures_df.empty:
        # Filtres
        col1, col2, col3 = st.columns(3)
        
        with col1:
            clients_list = ['Tous'] + list(factures_df['client_nom'].unique())
            filter_client = st.selectbox("Filtrer par client", clients_list)
        
        with col2:
            statuts_list = ['Tous'] + list(factures_df['statut'].unique())
            filter_statut = st.selectbox("Filtrer par statut", statuts_list)
        
        with col3:
            filter_date_debut = st.date_input("Date début", value=None, key="filter_date_debut")
            filter_date_fin = st.date_input("Date fin", value=None, key="filter_date_fin")
        
        # Appliquer les filtres
        filtered_df = factures_df.copy()
        
        if filter_client != 'Tous':
            filtered_df = filtered_df[filtered_df['client_nom'] == filter_client]
        
        if filter_statut != 'Tous':
            filtered_df = filtered_df[filtered_df['statut'] == filter_statut]
        
        if filter_date_debut:
            filtered_df = filtered_df[pd.to_datetime(filtered_df['date_emission']) >= pd.Timestamp(filter_date_debut)]
        
        if filter_date_fin:
            filtered_df = filtered_df[pd.to_datetime(filtered_df['date_emission']) <= pd.Timestamp(filter_date_fin)]
        
        if 'date_emission' in filtered_df.columns:
            filtered_df = filtered_df.sort_values('date_emission', ascending=False)
        
        # Indicateurs
        col1, col2, col3 = st.columns(3)
        
        with col1:
            ca_total = filtered_df['montant'].sum() if 'montant' in filtered_df.columns else 0
            st.metric("💰 CA Total", f"{ca_total:.2f} €")
        
        with col2:
            nb_factures = len(filtered_df)
            st.metric("📄 Nombre de factures", nb_factures)
        
        with col3:
            nb_payees = len(filtered_df[filtered_df['statut'] == 'payee']) if 'statut' in filtered_df.columns else 0
            st.metric("✅ Factures payées", nb_payees)
        
        # Tableau
        if not filtered_df.empty:
            display_columns = ['numero', 'date_emission', 'client_nom', 'description', 'montant', 'statut']
            display_df = filtered_df[display_columns].copy()
            
            if 'date_emission' in display_df.columns:
                display_df['date_emission'] = pd.to_datetime(display_df['date_emission']).dt.strftime('%d/%m/%Y')
            
            if 'montant' in display_df.columns:
                display_df['montant'] = display_df['montant'].apply(lambda x: f"{x:.2f} €")
            
            display_df.columns = ['Numéro', 'Date', 'Client', 'Description', 'Montant', 'Statut']

            # Réinitialiser l'index pour le cacher et afficher le DataFrame
            st.dataframe(display_df.reset_index(drop=True))
            
            # Actions
            st.subheader("📋 Actions")
            
            selected_numero = st.selectbox(
                "Sélectionner une facture",
                options=['-- Choisir --'] + list(filtered_df['numero'].unique()),
                key="action_numero"
            )
            
            if selected_numero != '-- Choisir --':
                facture = filtered_df[filtered_df['numero'] == selected_numero].iloc[0]
                is_locked = facture.get('locked', False) == True or facture.get('locked', False) == 'TRUE'
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    pdf_path = Path(facture.get('chemin_pdf', ''))
                    if pdf_path.exists() or (Path(__file__).parent.parent / pdf_path).exists():
                        full_path = (Path(__file__).parent.parent / pdf_path) if not pdf_path.is_absolute() else pdf_path
                        if full_path.exists():
                            with open(full_path, 'rb') as f:
                                st.download_button(
                                    "📥 Télécharger PDF",
                                    f.read(),
                                    file_name=pdf_path.name,
                                    mime='application/pdf'
                                )
                
                with col2:
                    if not is_locked:
                        current_statut = facture.get('statut', 'brouillon')
                        new_statut = st.selectbox(
                            "Statut",
                            options=['brouillon', 'payee'],
                            index=0 if current_statut == 'brouillon' else 1,
                            key="change_statut"
                        )
                        
                        if st.button("💾 Mettre à jour"):
                            factures_df.loc[factures_df['numero'] == selected_numero, 'statut'] = new_statut
                            excel_path = get_excel_path()
                            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                config_df = pd.read_excel(excel_path, sheet_name='CONFIG')
                                config_df.to_excel(writer, sheet_name='CONFIG', index=False)
                                factures_df.to_excel(writer, sheet_name='FACTURES', index=False)
                            st.success("Statut mis à jour !")
                            st.experimental_rerun()
                    else:
                        st.info("🔒 Facture verrouillée")
                
                with col3:
                    if is_locked:
                        if st.button("🔓 Déverrouiller"):
                            factures_df.loc[factures_df['numero'] == selected_numero, 'locked'] = False
                            excel_path = get_excel_path()
                            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                config_df = pd.read_excel(excel_path, sheet_name='CONFIG')
                                config_df.to_excel(writer, sheet_name='CONFIG', index=False)
                                factures_df.to_excel(writer, sheet_name='FACTURES', index=False)
                            st.success("Facture déverrouillée !")
                            st.experimental_rerun()
                    else:
                        if st.button("🔒 Verrouiller"):
                            factures_df.loc[factures_df['numero'] == selected_numero, 'locked'] = True
                            factures_df.loc[factures_df['numero'] == selected_numero, 'statut'] = 'payee'
                            excel_path = get_excel_path()
                            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                config_df = pd.read_excel(excel_path, sheet_name='CONFIG')
                                config_df.to_excel(writer, sheet_name='CONFIG', index=False)
                                factures_df.to_excel(writer, sheet_name='FACTURES', index=False)
                            st.success("Facture verrouillée !")
                            st.experimental_rerun()
        else:
            st.info("Aucune facture ne correspond aux filtres sélectionnés.")
    else:
        st.info("📝 Aucune facture pour le moment. Créez votre première facture ci-dessus !")
