"""
Page Scraping - Google Maps
Extraction des artisans depuis Google Maps
"""
import streamlit as st
import sys
import json
import time
from pathlib import Path
from datetime import datetime
import pandas as pd
import requests
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
# ✅ Plus besoin de threading/ThreadPoolExecutor - GitHub Actions uniquement

# Configuration de la page
st.set_page_config(page_title="Scraping Google Maps", page_icon="🔍", layout="wide")

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from whatsapp_database.queries import ajouter_artisan, get_statistiques
from whatsapp_database.models import init_database

# ✅ Initialiser la base de données au démarrage de la page (ajoute les nouvelles colonnes si nécessaire)
try:
    init_database()
except Exception as e:
    # Ne pas bloquer si erreur, mais logger
    import logging
    logging.warning(f"Erreur initialisation BDD: {e}")

# ✅ Import des fonctions de tracking (avec fallback si elles n'existent pas)
try:
    from whatsapp_database.queries import is_already_scraped, get_scraping_history, mark_scraping_done
except ImportError:
    # Si les fonctions n'existent pas encore, créer des stubs
    def is_already_scraped(metier: str, departement: str, ville: str) -> bool:
        return False
    
    def get_scraping_history(metier: str = None, departement: str = None):
        return []
    
    def mark_scraping_done(metier: str, departement: str, ville: str, results_count: int = 0):
        pass
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Charger les villes par département
try:
    with open(Path(__file__).parent.parent.parent / "data" / "villes_par_departement.json", 'r', encoding='utf-8') as f:
        villes_par_dept = json.load(f)
except:
    villes_par_dept = {}

# ✅ Fonction pour récupérer les communes depuis data.gouv.fr
def get_communes_from_api(departement: str, min_population: int = 0, max_population: int = 50000):
    """Récupère les communes d'un département depuis l'API data.gouv.fr avec coordonnées GPS"""
    try:
        url = f"https://geo.api.gouv.fr/departements/{departement}/communes"
        params = {
            "fields": "nom,code,codesPostaux,population,centre",
            "format": "json"
        }
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 200:
            communes = response.json()
            # Filtrer par population
            filtered = []
            for c in communes:
                pop = c.get('population', 0)
                if min_population <= pop <= max_population:
                    centre = c.get('centre', {})
                    filtered.append({
                        'nom': c['nom'],
                        'code': c['code'],
                        'code_postal': c.get('codesPostaux', [c.get('code', '')])[0] if c.get('codesPostaux') else c.get('code', ''),
                        'population': pop,
                        'latitude': centre.get('coordinates', [None, None])[1] if centre else None,
                        'longitude': centre.get('coordinates', [None, None])[0] if centre else None
                    })
            # Trier par population (croissant)
            filtered.sort(key=lambda x: x['population'])
            return filtered
    except Exception as e:
        logger.error(f"Erreur API communes: {e}")
    return []

st.title("🔍 Scraping Google Maps - Artisans")

# Stats actuelles
stats = get_statistiques()
col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
with col_stat1:
    st.metric("Total artisans", f"{stats.get('total', 0):,}")
with col_stat2:
    st.metric("Avec téléphone", f"{stats.get('avec_telephone', 0):,}")
with col_stat3:
    st.metric("Avec site web", f"{stats.get('avec_site_web', 0):,}")
with col_stat4:
    st.metric("Sans site web", f"{stats.get('sans_site_web', 0):,}")

st.markdown("---")

# Configuration du scraping
st.subheader("⚙️ Configuration")

col_config1, col_config2 = st.columns(2)

with col_config1:
    # ✅ Multi-select pour les métiers
    metiers_options = ["plombier", "électricien", "chauffagiste", "menuisier", "peintre", "maçon", "couvreur", "carreleur"]
    metiers = st.multiselect(
        "Type(s) d'artisan(s)",
        options=metiers_options,
        default=["plombier"],
        help="Sélectionnez un ou plusieurs types d'artisans à rechercher"
    )
    if not metiers:
        st.warning("⚠️ Veuillez sélectionner au moins un métier")
        metier = "plombier"
    else:
        metier = metiers[0]

with col_config2:
    # Liste des départements français
    departements_liste = [
        "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
        "11", "12", "13", "14", "15", "16", "17", "18", "19", "21",
        "22", "23", "24", "25", "26", "27", "28", "29", "2A", "2B",
        "30", "31", "32", "33", "34", "35", "36", "37", "38", "39",
        "40", "41", "42", "43", "44", "45", "46", "47", "48", "49",
        "50", "51", "52", "53", "54", "55", "56", "57", "58", "59",
        "60", "61", "62", "63", "64", "65", "66", "67", "68", "69",
        "70", "71", "72", "73", "74", "75", "76", "77", "78", "79",
        "80", "81", "82", "83", "84", "85", "86", "87", "88", "89",
        "90", "91", "92", "93", "94", "95"
    ]
    # ✅ Multi-select pour les départements
    departements = st.multiselect(
        "Département(s)",
        options=departements_liste,
        default=["77"],
        help="Sélectionnez un ou plusieurs départements à scraper"
    )
    if not departements:
        st.warning("⚠️ Veuillez sélectionner au moins un département")
        departement = "77"
    else:
        departement = departements[0]

col_config3, col_config4 = st.columns(2)

with col_config3:
    max_results = st.slider(
        "Nombre max de résultats",
        min_value=10,
        max_value=200,
        value=50,
        step=10,
        help="Nombre maximum d'établissements à scraper par ville"
    )

with col_config4:
    headless = st.checkbox(
        "Mode headless (navigateur invisible)",
        value=True,
        help="Mode headless activé par défaut (plus rapide). Décochez pour voir le navigateur."
    )

# ✅ Charger la configuration GitHub si elle existe
github_config_file = Path(__file__).parent.parent.parent / "config" / "github_config.json"
github_token_default = ""
github_repo_default = ""

try:
    if github_config_file.exists():
        with open(github_config_file, 'r', encoding='utf-8') as f:
            github_config = json.load(f)
            github_token_default = github_config.get('github_token', '')
            github_repo_default = github_config.get('github_repo', '')
except:
    pass

# ✅ GitHub Actions est maintenant la SEULE option disponible
# Forcer GitHub Actions à True
use_github_actions = True
st.session_state.use_github_actions = True

# ✅ Utiliser les valeurs du fichier de config automatiquement (pas de champs visibles)
github_token = github_token_default
github_repo = github_repo_default

if not github_token or not github_repo:
    st.error("⚠️ Configuration GitHub manquante. Vérifiez que config/github_config.json existe avec token et repo.")

# ✅ Section : Gestion des workflows GitHub Actions (VISIBLE EN HAUT, DÈS LE DÉMARRAGE)
# ✅ TOUJOURS AFFICHÉE - même si pas de token (pour montrer qu'il faut configurer)
st.markdown("---")
st.subheader("⚙️ Gestion des Workflows GitHub Actions")

# ✅ FIX CRITIQUE : Définir les fonctions AVANT leur utilisation
def list_github_workflows(token, repo):
    """Liste tous les workflows GitHub Actions en cours"""
    try:
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        
        workflows = []
        
        # ✅ FIX : Faire deux appels séparés car l'API ne supporte pas "in_progress,queued" dans un seul paramètre
        # Récupérer les runs "in_progress"
        runs_url_in_progress = f"https://api.github.com/repos/{repo}/actions/runs?status=in_progress&per_page=100"
        response = requests.get(runs_url_in_progress, headers=headers)
        if response.status_code == 200:
            runs_data = response.json()
            for run in runs_data.get('workflow_runs', []):
                workflows.append({
                    'id': run.get('id'),
                    'run_number': run.get('run_number'),
                    'status': run.get('status'),
                    'conclusion': run.get('conclusion'),
                    'created_at': run.get('created_at'),
                    'updated_at': run.get('updated_at'),
                    'head_branch': run.get('head_branch'),
                    'workflow_id': run.get('workflow_id'),
                    'html_url': run.get('html_url')
                })
        
        # Récupérer les runs "queued"
        runs_url_queued = f"https://api.github.com/repos/{repo}/actions/runs?status=queued&per_page=100"
        response = requests.get(runs_url_queued, headers=headers)
        if response.status_code == 200:
            runs_data = response.json()
            for run in runs_data.get('workflow_runs', []):
                # Éviter les doublons
                if not any(w['id'] == run.get('id') for w in workflows):
                    workflows.append({
                        'id': run.get('id'),
                        'run_number': run.get('run_number'),
                        'status': run.get('status'),
                        'conclusion': run.get('conclusion'),
                        'created_at': run.get('created_at'),
                        'updated_at': run.get('updated_at'),
                        'head_branch': run.get('head_branch'),
                        'workflow_id': run.get('workflow_id'),
                        'html_url': run.get('html_url')
                    })
        
        # Trier par date de création (plus récent en premier)
        workflows.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        
        return workflows
    except Exception as e:
        logger.error(f"Erreur liste workflows: {e}")
        return []

def cancel_github_workflow(token, repo, run_id):
    """Annule un workflow GitHub Actions spécifique"""
    try:
        cancel_url = f"https://api.github.com/repos/{repo}/actions/runs/{run_id}/cancel"
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        
        response = requests.post(cancel_url, headers=headers)
        return response.status_code == 202
    except Exception as e:
        logger.error(f"Erreur annulation workflow {run_id}: {e}")
        return False

def cancel_all_github_workflows(token, repo):
    """Annule tous les workflows GitHub Actions en cours (in_progress et queued)"""
    try:
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        
        cancelled_count = 0
        total_runs = 0
        
        # Récupérer les runs "in_progress"
        url_in_progress = f"https://api.github.com/repos/{repo}/actions/runs?status=in_progress&per_page=100"
        response = requests.get(url_in_progress, headers=headers)
        if response.status_code == 200:
            runs_data = response.json()
            for run in runs_data.get('workflow_runs', []):
                total_runs += 1
                if cancel_github_workflow(token, repo, run.get('id')):
                    cancelled_count += 1
        
        # Récupérer les runs "queued"
        url_queued = f"https://api.github.com/repos/{repo}/actions/runs?status=queued&per_page=100"
        response = requests.get(url_queued, headers=headers)
        if response.status_code == 200:
            runs_data = response.json()
            for run in runs_data.get('workflow_runs', []):
                total_runs += 1
                if cancel_github_workflow(token, repo, run.get('id')):
                    cancelled_count += 1
        
        if total_runs == 0:
            return True, "Aucun workflow en cours à annuler"
        elif cancelled_count == total_runs:
            return True, f"✅ {cancelled_count} workflow(s) annulé(s) avec succès"
        else:
            return False, f"⚠️ {cancelled_count}/{total_runs} workflow(s) annulé(s)"
    except Exception as e:
        logger.error(f"Erreur annulation workflows: {e}")
        return False, f"Erreur: {str(e)}"

def download_github_artifact(token, repo, run_id):
    """Télécharge l'artifact depuis GitHub Actions"""
    try:
        # Récupérer la liste des artifacts
        url = f"https://api.github.com/repos/{repo}/actions/runs/{run_id}/artifacts"
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            artifacts = response.json().get('artifacts', [])
            for artifact in artifacts:
                if artifact.get('name') == 'scraping-results':
                    # Télécharger l'artifact
                    download_url = artifact.get('archive_download_url')
                    if download_url:
                        download_response = requests.get(download_url, headers=headers)
                        if download_response.status_code == 200:
                            # Sauvegarder le zip
                            import zipfile
                            import io
                            data_dir = Path(__file__).parent.parent.parent / "data"
                            zip_path = data_dir / "github_artifact.zip"
                            with open(zip_path, 'wb') as f:
                                f.write(download_response.content)
                            
                            # Extraire le JSON
                            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                                zip_ref.extractall(data_dir)
                            
                            # Lire les fichiers
                            results_file = data_dir / "scraping_results_github_actions.json"
                            status_file = data_dir / "github_actions_status.json"
                            
                            result_data = None
                            status_data = None
                            
                            if results_file.exists():
                                with open(results_file, 'r', encoding='utf-8') as f:
                                    result_data = json.load(f)
                            
                            if status_file.exists():
                                with open(status_file, 'r', encoding='utf-8') as f:
                                    status_data = json.load(f)
                            
                            # Nettoyer
                            zip_path.unlink()
                            
                            # ✅ Retourner dans le format attendu (compatibilité)
                            if result_data and isinstance(result_data, dict) and 'results' in result_data:
                                return result_data  # Format: {'results': [...], 'total_results': ...}
                            elif result_data and isinstance(result_data, list):
                                return {'results': result_data, 'total_results': len(result_data)}
                            else:
                                return {'results': [], 'total_results': 0, 'status': status_data}
            return None
        return None
    except Exception as e:
        logger.error(f"Erreur téléchargement artifact: {e}")
        return None

if github_token and github_repo:
    # ✅ Bouton Rafraîchir TOUJOURS visible EN HAUT pour forcer la mise à jour
    col_refresh_top1, col_refresh_top2 = st.columns([1, 1])
    with col_refresh_top1:
        if st.button("🔄 Rafraîchir les workflows", key="refresh_workflows_top", help="Afficher tous les workflows et leurs statistiques"):
            # Forcer la mise à jour en réinitialisant le cache de session
            if 'workflows_last_refresh' in st.session_state:
                del st.session_state.workflows_last_refresh
            # Forcer le rerun immédiatement
            st.experimental_rerun()

    # Lister les workflows en cours
    try:
        workflows_en_cours = list_github_workflows(github_token, github_repo)
    except Exception as e:
        logger.error(f"Erreur récupération workflows: {e}")
        workflows_en_cours = []
    
    # ✅ Affichage simplifié - pas de détails individuels qui apparaissent/disparaissent
    if workflows_en_cours:
        st.info(f"🟢 **{len(workflows_en_cours)} workflow(s) en cours** - Utilisez le bouton 'Rafraîchir les workflows' ci-dessus pour voir les détails")
    
    # ✅ Bouton supprimé - déjà présent dans la page Base de données
    
    # ✅ Afficher les statistiques pour chaque workflow (même terminés)
    # Récupérer TOUS les workflows (pas seulement in_progress/queued)
    try:
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {github_token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        # ✅ Récupérer les 4 derniers workflows (tous statuts)
        all_workflows_url = f"https://api.github.com/repos/{github_repo}/actions/runs?per_page=4"
        response = requests.get(all_workflows_url, headers=headers)
        all_workflows = []
        if response.status_code == 200:
            runs_data = response.json()
            for run in runs_data.get('workflow_runs', []):
                all_workflows.append({
                    'id': run.get('id'),
                    'run_number': run.get('run_number'),
                    'status': run.get('status'),
                    'conclusion': run.get('conclusion'),
                    'created_at': run.get('created_at'),
                    'html_url': run.get('html_url')
                })
            all_workflows.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    except:
        all_workflows = workflows_en_cours
    
    if all_workflows:
        st.markdown("### 📊 Statistiques par workflow")
        from whatsapp_database.queries import get_artisans
        
        # ✅ Afficher en grille 2 colonnes (4 workflows max)
        workflows_to_show = all_workflows[:4]  # Limiter aux 4 derniers
        
        # Créer des paires de workflows pour affichage en 2 colonnes
        for i in range(0, len(workflows_to_show), 2):
            cols = st.columns(2)
            for j, col in enumerate(cols):
                if i + j < len(workflows_to_show):
                    workflow = workflows_to_show[i + j]
                    status_emoji = "🟢" if workflow['status'] == 'in_progress' else "🟡" if workflow['status'] == 'queued' else "🔵"
                    conclusion_emoji = "✅" if workflow.get('conclusion') == 'success' else "❌" if workflow.get('conclusion') == 'failure' else "⏹️" if workflow.get('conclusion') == 'cancelled' else ""
                    
                    with col:
                        with st.expander(f"{status_emoji} {conclusion_emoji} Workflow #{workflow['run_number']} - {workflow['status']} ({workflow['created_at'][:19].replace('T', ' ')})"):
                            # Récupérer les artisans scrapés depuis le début de ce workflow
                            workflow_start = workflow['created_at']
                            
                            try:
                                # Récupérer tous les artisans
                                all_artisans = get_artisans(limit=10000)
                                
                                # ✅ Normaliser le format de workflow_start (GitHub API format: "2025-11-28T16:56:17Z")
                                # Convertir en format comparable (sans Z, avec espace au lieu de T)
                                workflow_start_normalized = workflow_start.replace('T', ' ').replace('Z', '').split('.')[0]
                                
                                # Filtrer ceux créés après le début du workflow
                                workflow_artisans = []
                                for a in all_artisans:
                                    created_at = a.get('created_at')
                                    if created_at:
                                        # Normaliser created_at (peut être ISO ou SQLite format)
                                        created_at_normalized = str(created_at).replace('T', ' ').replace('Z', '').split('.')[0]
                                        # Comparaison de chaînes ISO normalisées (format: "YYYY-MM-DD HH:MM:SS")
                                        if created_at_normalized >= workflow_start_normalized:
                                            workflow_artisans.append(a)
                                
                                # ✅ Debug: Afficher le nombre total d'artisans et ceux filtrés
                                if len(all_artisans) > 0 and len(workflow_artisans) == 0:
                                    # Si on a des artisans mais aucun ne correspond, c'est peut-être un problème de format
                                    # Afficher les 3 derniers artisans pour debug
                                    st.caption(f"🔍 Debug: {len(all_artisans)} artisans totaux, workflow_start: {workflow_start_normalized}")
                                
                                if workflow_artisans:
                                    total = len(workflow_artisans)
                                    avec_tel = len([a for a in workflow_artisans if a.get('telephone')])
                                    avec_site = len([a for a in workflow_artisans if a.get('site_web')])
                                    sans_site = total - avec_site
                                    
                                    # ✅ Stats simples sans colonnes imbriquées
                                    st.markdown(f"**📊 Scrapés:** {total} | **📞 Avec téléphone:** {avec_tel} | **🌐 Avec site web:** {avec_site} | **⭐ SANS site:** {sans_site}")
                                else:
                                    st.info("⏳ Aucun résultat encore pour ce workflow")
                            except Exception as e:
                                st.error(f"Erreur calcul stats: {e}")
else:
    st.warning("⚠️ Configuration GitHub manquante. La gestion des workflows nécessite un token et un repository configurés.")

st.markdown("---")

# ✅ SECTION : Carte de scraping par métier (basée sur la BDD locale)
st.markdown("### 🗺️ Carte des scrapings par métier")
st.caption("Visualisez les départements scrapés pour chaque métier. La taille des points est proportionnelle au nombre d'artisans.")

# Récupérer la liste des métiers depuis la BDD
from whatsapp_database.queries import get_artisans
all_artisans = get_artisans(limit=10000)
metiers_bdd = sorted(list(set([a.get('type_artisan') for a in all_artisans if a.get('type_artisan')])))

col_map1, col_map2 = st.columns([1, 3])
with col_map1:
    if metiers_bdd:
        metier_carte = st.selectbox("Sélectionner un métier", ["Tous"] + metiers_bdd, key="map_metier_selection")
    else:
        metier_carte = "Tous"
        st.info("ℹ️ Aucun métier dans la BDD")

# Importer et utiliser la fonction de carte
try:
    from whatsapp_app.utils.map_utils import create_scraping_map_by_job
    from streamlit_folium import st_folium
    
    carte = create_scraping_map_by_job(metier_carte if metier_carte != "Tous" else None)
    
    if carte:
        with col_map2:
            st_folium(carte, width=None, height=500, returned_objects=[])
    else:
        with col_map2:
            st.info("ℹ️ Aucune donnée de scraping pour ce métier. Lancez un scraping pour voir la carte se remplir.")
except ImportError as e:
    st.warning(f"⚠️ Erreur import map_utils: {e}")
    st.info("💡 Installez les dépendances: `pip install folium streamlit-folium`")
except Exception as e:
    st.error(f"❌ Erreur création carte: {e}")
    import traceback
    st.code(traceback.format_exc())

# ✅ SECTION : Tableau de bord stratégique de recherche
try:
    from whatsapp_database.scraping_analytics import (
        get_metier_statistics, get_departement_statistics,
        get_coverage_metrics, get_session_statistics,
        get_priority_suggestions, generate_research_report
    )
    ANALYTICS_AVAILABLE = True
except ImportError as e:
    ANALYTICS_AVAILABLE = False
    # Debug: afficher l'erreur en mode développement
    import traceback
    # st.warning(f"Module analytics non disponible: {e}")  # Commenté pour ne pas polluer l'UI

# ✅ SECTION : Historique des scrapings par département (masqué si vide)
from whatsapp_database.queries import get_scraping_history
historique = get_scraping_history()

# ✅ Initialiser departements_stats pour éviter l'erreur
departements_stats = {}

if historique:
    col_title, col_filter = st.columns([2, 1])
    with col_title:
        st.markdown("### 📊 Historique des scrapings par département")
    with col_filter:
        # Sélection du métier pour filtrer
        metiers_disponibles = list(set([h['metier'] for h in historique if h.get('metier')]))
        if metiers_disponibles:
            metier_selectionne = st.selectbox("Métier", ["Tous"] + sorted(metiers_disponibles), key="tracking_metier", label_visibility="visible")
        else:
            metier_selectionne = "Tous"
    
    # Filtrer par métier si spécifié
    if metier_selectionne != "Tous":
        historique = [h for h in historique if h.get('metier') == metier_selectionne]
    
    # Calculer les statistiques par département
    for entry in historique:
        dept = entry.get('departement', '')
        metier = entry.get('metier', '')
        results_count = entry.get('results_count', 0)
        
        if not dept:
            continue
        
        key = f"{metier}_{dept}"
        if key not in departements_stats:
            departements_stats[key] = {
                'departement': dept,
                'metier': metier,
                'total_results': 0,
                'villes_scrapees': 0,
                'max_results': 0
            }
        
        departements_stats[key]['total_results'] += results_count
        departements_stats[key]['villes_scrapees'] += 1
        departements_stats[key]['max_results'] = max(departements_stats[key]['max_results'], results_count)

# Coordonnées approximatives des départements français (centres)
dept_coords = {
    '01': (46.2, 5.2), '02': (49.4, 3.4), '03': (46.3, 3.1), '04': (44.1, 6.1), '05': (44.7, 6.1),
    '06': (43.7, 7.3), '07': (44.5, 4.4), '08': (49.8, 4.7), '09': (43.0, 1.6), '10': (48.3, 4.1),
    '11': (43.2, 2.4), '12': (44.3, 2.6), '13': (43.3, 5.4), '14': (49.2, -0.4), '15': (45.0, 2.4),
    '16': (45.6, 0.2), '17': (46.2, -1.2), '18': (47.1, 2.4), '19': (45.3, 1.8), '21': (47.3, 5.0),
    '22': (48.5, -2.8), '23': (46.2, 1.9), '24': (45.2, 0.7), '25': (47.2, 6.0), '26': (44.9, 4.9),
    '27': (49.1, 1.1), '28': (48.4, 1.5), '29': (48.4, -4.5), '30': (44.1, 4.1), '31': (43.6, 1.4),
    '32': (43.6, 0.6), '33': (44.8, -0.6), '34': (43.6, 3.9), '35': (48.1, -1.7), '36': (46.8, 1.7),
    '37': (47.4, 0.7), '38': (45.2, 5.7), '39': (46.7, 5.6), '40': (43.9, -0.5), '41': (47.6, 1.3),
    '42': (45.4, 4.4), '43': (45.0, 3.9), '44': (47.2, -1.6), '45': (47.9, 1.9), '46': (44.4, 1.4),
    '47': (44.2, 0.6), '48': (44.5, 3.5), '49': (47.5, -0.6), '50': (49.1, -1.1), '51': (49.3, 4.0),
    '52': (48.1, 5.1), '53': (48.1, -0.8), '54': (48.7, 6.2), '55': (49.1, 5.4), '56': (47.7, -2.8),
    '57': (49.1, 6.2), '58': (47.0, 3.4), '59': (50.6, 3.1), '60': (49.4, 2.8), '61': (48.4, 0.1),
    '62': (50.3, 2.8), '63': (45.8, 3.1), '64': (43.3, -0.4), '65': (43.2, 0.1), '66': (42.7, 2.9),
    '67': (48.6, 7.8), '68': (47.7, 7.3), '69': (45.8, 4.8), '70': (47.6, 6.2), '71': (46.8, 4.9),
    '72': (48.0, 0.2), '73': (45.6, 5.9), '74': (46.0, 6.1), '75': (48.9, 2.3), '76': (49.4, 1.1),
    '77': (48.6, 2.7), '78': (48.8, 2.1), '79': (46.3, -0.5), '80': (49.9, 2.3), '81': (43.6, 2.1),
    '82': (44.0, 1.4), '83': (43.1, 6.0), '84': (44.0, 5.0), '85': (46.7, -1.4), '86': (46.6, 0.3),
    '87': (45.8, 1.3), '88': (48.2, 6.5), '89': (47.8, 3.6), '90': (47.6, 6.9), '91': (48.6, 2.3),
    '92': (48.9, 2.2), '93': (48.9, 2.4), '94': (48.8, 2.4), '95': (49.1, 2.3)
}

# Créer la carte
if departements_stats:
    import folium
    from streamlit_folium import st_folium
    
    # Centre de la France
    m = folium.Map(location=[46.6, 2.2], zoom_start=6)
    
    # Seuils pour déterminer le statut
    SEUIL_FAIT = 5  # Au moins 5 résultats = "fait"
    SEUIL_PARTIEL = 1  # Entre 1 et 4 = "partiellement fait"
    
    for key, stats in departements_stats.items():
        dept = stats['departement']
        metier = stats['metier']
        total_results = stats['total_results']
        villes_scrapees = stats['villes_scrapees']
        max_results = stats['max_results']
        
        if dept in dept_coords:
            lat, lon = dept_coords[dept]
            
            # Déterminer le statut et la couleur
            if max_results >= SEUIL_FAIT:
                couleur = 'green'
                statut = '✅ Fait'
                taille = 15
            elif max_results >= SEUIL_PARTIEL:
                couleur = 'orange'
                statut = '⚠️ Partiellement fait'
                taille = 10
            else:
                couleur = 'gray'
                statut = '❌ Non fait'
                taille = 5
            
            # Popup avec informations
            popup_text = f"""
            <b>Département {dept}</b><br>
            <b>Métier:</b> {metier}<br>
            <b>Statut:</b> {statut}<br>
            <b>Villes scrapées:</b> {villes_scrapees}<br>
            <b>Total résultats:</b> {total_results}<br>
            <b>Max par ville:</b> {max_results}
            """
            
            folium.CircleMarker(
                location=[lat, lon],
                radius=taille,
                popup=folium.Popup(popup_text, max_width=300),
                color='black',
                weight=2,
                fillColor=couleur,
                fillOpacity=0.8,
                tooltip=f"Dépt {dept}: {statut} ({villes_scrapees} villes, {total_results} résultats)"
            ).add_to(m)
    
    # ✅ AUSSI afficher les villes individuelles scrapées (petits points bleus)
    # Récupérer les coordonnées des villes depuis l'API ou depuis les artisans scrapés
    from whatsapp_database.queries import get_artisans
    artisans = get_artisans(limit=10000)
    
    # Grouper les villes scrapées par métier/département
    villes_scrapees_coords = {}
    for artisan in artisans:
        if metier_selectionne != "Tous" and artisan.get('type_artisan') != metier_selectionne:
            continue
        dept_art = artisan.get('departement', '')
        ville_art = artisan.get('ville', '')
        if dept_art and ville_art:
            key = f"{dept_art}_{ville_art}"
            if key not in villes_scrapees_coords:
                # Essayer de récupérer les coordonnées depuis l'API
                try:
                    communes = get_communes_from_api(dept_art, 0, 1000000)  # Toutes les communes
                    for commune in communes:
                        if commune['nom'].lower() == ville_art.lower():
                            villes_scrapees_coords[key] = {
                                'lat': commune.get('latitude'),
                                'lon': commune.get('longitude'),
                                'ville': ville_art,
                                'dept': dept_art
                            }
                            break
                except:
                    pass
    
    # Ajouter les villes sur la carte (petits points bleus)
    for key, coords in villes_scrapees_coords.items():
        if coords.get('lat') and coords.get('lon'):
            folium.CircleMarker(
                location=[coords['lat'], coords['lon']],
                radius=3,
                popup=folium.Popup(f"<b>{coords['ville']}</b><br>Département: {coords['dept']}", max_width=200),
                color='blue',
                weight=1,
                fillColor='blue',
                fillOpacity=0.6,
                tooltip=f"{coords['ville']} ({coords['dept']})"
            ).add_to(m)
    
    # Légende
    legend_html = '''
    <div style="position: fixed; 
                bottom: 50px; left: 50px; width: 200px; height: 120px; 
                background-color: white; z-index:9999; font-size:14px;
                border:2px solid grey; border-radius:5px; padding: 10px">
    <b>Légende</b><br>
    <span style="color:green">●</span> Fait (≥5 résultats)<br>
    <span style="color:orange">●</span> Partiel (1-4 résultats)<br>
    <span style="color:gray">●</span> Non fait (0 résultat)<br>
    <span style="color:blue">●</span> Villes scrapées
    </div>
    '''
    m.get_root().html.add_child(folium.Element(legend_html))
    
    # ✅ Afficher la carte EN GRAND
    st_folium(m, width=None, height=600, returned_objects=[])
    
    # Statistiques résumées
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    with col_stat1:
        total_depts = len(departements_stats)
        st.metric("Départements scrapés", total_depts)
    with col_stat2:
        depts_faits = len([s for s in departements_stats.values() if s['max_results'] >= SEUIL_FAIT])
        st.metric("Départements complets", depts_faits)
    with col_stat3:
        total_villes = sum([s['villes_scrapees'] for s in departements_stats.values()])
        st.metric("Villes scrapées", total_villes)
else:
    st.empty()  # Pas de message si aucun scraping - affichage plus compact

# ✅ NOUVELLE SECTION : Tableau de bord stratégique de recherche
st.markdown("---")
st.markdown("### 📊 Tableau de bord stratégique de recherche")

if not ANALYTICS_AVAILABLE:
    st.warning("⚠️ Le module d'analytics n'est pas disponible. Vérifiez que le fichier `whatsapp_database/scraping_analytics.py` existe.")
    st.stop()

if not historique or len(historique) == 0:
    st.info("ℹ️ Aucun historique de scraping trouvé dans la base de données. Les statistiques apparaîtront après vos premiers scrapings qui seront enregistrés dans `scraping_history`.")
    st.caption("💡 Note: Les scrapings effectués via GitHub Actions doivent être marqués dans la table `scraping_history` pour apparaître ici.")
    
    # Bouton pour forcer la réinitialisation de la BDD (utile si les colonnes manquent)
    if st.button("🔄 Réinitialiser la structure de la base de données", help="Ajoute les nouvelles colonnes à la table scraping_history si elles manquent"):
        try:
            from whatsapp_database.models import init_database
            init_database()
            st.success("✅ Base de données réinitialisée ! Rafraîchissez la page.")
            try:
                st.rerun()
            except:
                st.experimental_rerun()
        except Exception as e:
            st.error(f"❌ Erreur: {e}")
else:
    # Tabs pour organiser les différentes vues
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🎯 Vue d'ensemble", 
        "📈 Par métier", 
        "🗺️ Par département",
        "📅 Sessions récentes",
        "💡 Suggestions"
    ])
    
    with tab1:
        st.markdown("#### Vue d'ensemble globale")
        
        # Statistiques globales
        try:
            stats_metiers = get_metier_statistics()
            stats_departements = get_departement_statistics()
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Métiers scrapés", len(stats_metiers))
            with col2:
                st.metric("Départements couverts", len(stats_departements))
            with col3:
                total_artisans = sum(s['artisans_trouves'] for s in stats_metiers)
                st.metric("Artisans trouvés", f"{total_artisans:,}")
            with col4:
                total_villes = sum(s['villes_scrapees'] for s in stats_metiers)
                st.metric("Villes scrapées", f"{total_villes:,}")
            
            # Graphique des métiers les plus scrapés
            if stats_metiers:
                st.markdown("#### Top métiers par nombre d'artisans trouvés")
                df_metiers = pd.DataFrame(stats_metiers[:10])
                st.bar_chart(df_metiers.set_index('metier')['artisans_trouves'])
                
        except Exception as e:
            st.warning(f"⚠️ Erreur calcul statistiques: {e}")
    
    with tab2:
        st.markdown("#### Statistiques par métier")
        
        try:
            stats_metiers = get_metier_statistics()
            
            if stats_metiers:
                # Tableau détaillé
                df = pd.DataFrame(stats_metiers)
                df['taux_couverture_moyen'] = df['taux_couverture_moyen'].apply(
                    lambda x: f"{x:.1f}%" if x is not None else "N/A"
                )
                
                st.dataframe(
                    df[['metier', 'departements_couverts', 'villes_scrapees', 
                        'artisans_trouves', 'taux_couverture_moyen']],
                    use_container_width=True,
                    hide_index=True
                )
                
                # Export
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    "📥 Télécharger en CSV",
                    csv,
                    f"stats_metiers_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv"
                )
            else:
                st.info("Aucune statistique disponible")
                
        except Exception as e:
            st.warning(f"⚠️ Erreur: {e}")
    
    with tab3:
        st.markdown("#### Statistiques par département")
        
        try:
            stats_departements = get_departement_statistics()
            
            if stats_departements:
                # Filtre par métier
                metiers_list = sorted(list(set([h.get('metier', '') for h in historique if h.get('metier')])))
                if metiers_list:
                    metier_filter = st.selectbox(
                        "Filtrer par métier",
                        ["Tous"] + metiers_list,
                        key="dept_metier_filter"
                    )
                    
                    if metier_filter != "Tous":
                        stats_departements = [s for s in stats_departements 
                                            if any(h.get('metier') == metier_filter 
                                                  for h in historique if h.get('departement') == s['departement'])]
                
                # Tableau
                df = pd.DataFrame(stats_departements)
                df['taux_couverture'] = df['taux_couverture'].apply(
                    lambda x: f"{x:.1f}%" if x is not None else "N/A"
                )
                
                st.dataframe(
                    df[['departement', 'metiers_scrapes', 'villes_scrapees', 
                        'villes_disponibles', 'taux_couverture', 'artisans_trouves']],
                    use_container_width=True,
                    hide_index=True
                )
                
                # Export
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    "📥 Télécharger en CSV",
                    csv,
                    f"stats_departements_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv"
                )
            else:
                st.info("Aucune statistique disponible")
                
        except Exception as e:
            st.warning(f"⚠️ Erreur: {e}")
    
    with tab4:
        st.markdown("#### Sessions de scraping récentes (30 derniers jours)")
        
        try:
            sessions = get_session_statistics(days=30)
            
            if sessions:
                df_sessions = pd.DataFrame(sessions)
                df_sessions['duree_moyenne'] = df_sessions['duree_moyenne'].apply(
                    lambda x: f"{x:.0f}s" if x and not pd.isna(x) else "N/A"
                )
                
                st.dataframe(
                    df_sessions,
                    use_container_width=True,
                    hide_index=True
                )
                
                # Graphique d'évolution
                if len(df_sessions) > 1:
                    st.markdown("#### Évolution du nombre d'artisans trouvés")
                    st.line_chart(df_sessions.set_index('date')['artisans_trouves'])
            else:
                st.info("Aucune session récente")
                
        except Exception as e:
            st.warning(f"⚠️ Erreur: {e}")
    
    with tab5:
        st.markdown("#### Suggestions de départements prioritaires")
        st.caption("Départements partiellement couverts nécessitant un complément de scraping")
        
        try:
            suggestions = get_priority_suggestions(limit=10)
            
            if suggestions:
                df_suggestions = pd.DataFrame(suggestions)
                df_suggestions['taux_couverture_actuel'] = df_suggestions['taux_couverture_actuel'].apply(
                    lambda x: f"{x:.1f}%" if x is not None else "N/A"
                )
                df_suggestions['priorite_score'] = df_suggestions['priorite_score'].apply(
                    lambda x: f"{x:.0f}"
                )
                
                st.dataframe(
                    df_suggestions[['departement', 'metier', 'taux_couverture_actuel', 
                                   'villes_manquantes', 'raison', 'priorite_score']],
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.success("✅ Tous les départements sont bien couverts !")
                
        except Exception as e:
            st.warning(f"⚠️ Erreur: {e}")
    
    # Section Rapport complet
    st.markdown("---")
    st.markdown("#### 📄 Générer un rapport complet")
    
    col_report1, col_report2 = st.columns(2)
    with col_report1:
        metier_report = st.selectbox(
            "Métier (optionnel)",
            ["Tous"] + sorted(list(set([h.get('metier', '') for h in historique if h.get('metier')]))),
            key="report_metier"
        )
    with col_report2:
        dept_report = st.selectbox(
            "Département (optionnel)",
            ["Tous"] + sorted(list(set([h.get('departement', '') for h in historique if h.get('departement')]))),
            key="report_dept"
        )
    
    if st.button("📊 Générer rapport", key="generate_report"):
        try:
            with st.spinner("Génération du rapport..."):
                report = generate_research_report(
                    metier=metier_report if metier_report != "Tous" else None,
                    departement=dept_report if dept_report != "Tous" else None
                )
                
                st.success("✅ Rapport généré avec succès !")
                
                # Afficher le résumé
                st.markdown("##### Résumé")
                stats_gen = report['statistiques_generales']
                col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                with col_r1:
                    st.metric("Scrapings", stats_gen['total_scrapings'])
                with col_r2:
                    st.metric("Artisans", f"{stats_gen['total_artisans_trouves']:,}")
                with col_r3:
                    st.metric("Villes", stats_gen['villes_scrapees'])
                with col_r4:
                    st.metric("Départements", stats_gen['departements_couverts'])
                
                # Exports
                col_exp1, col_exp2 = st.columns(2)
                
                with col_exp1:
                    # Export JSON
                    import json
                    report_json = json.dumps(report, ensure_ascii=False, indent=2)
                    st.download_button(
                        "📥 Télécharger rapport JSON",
                        report_json.encode('utf-8'),
                        f"rapport_recherche_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                        "application/json"
                    )
                
                with col_exp2:
                    # Export Excel si disponible
                    if EXCEL_AVAILABLE:
                        try:
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Résumé"
                            
                            # Style header
                            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF")
                            
                            # Feuille 1: Résumé
                            ws.append(["Rapport de recherche"])
                            ws.append(["Généré le", report['periode']['generated_at']])
                            ws.append([])
                            
                            stats = report['statistiques_generales']
                            ws.append(["Statistiques générales"])
                            ws.append(["Total scrapings", stats['total_scrapings']])
                            ws.append(["Artisans trouvés", stats['total_artisans_trouves']])
                            ws.append(["Villes scrapées", stats['villes_scrapees']])
                            ws.append(["Départements couverts", stats['departements_couverts']])
                            ws.append(["Métiers scrapés", stats['metiers_scrapes']])
                            ws.append(["Moyenne artisans/scraping", f"{stats['artisans_moyen_par_scraping']:.2f}"])
                            
                            # Feuille 2: Par métier
                            if report['statistiques_par_metier']:
                                ws2 = wb.create_sheet("Par métier")
                                df_metiers = pd.DataFrame(report['statistiques_par_metier'])
                                for r_idx, row in enumerate(df_metiers.itertuples(), start=1):
                                    for c_idx, value in enumerate(row[1:], start=1):
                                        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                                        if r_idx == 1:
                                            cell.fill = header_fill
                                            cell.font = header_font
                            
                            # Feuille 3: Par département
                            if report['statistiques_par_departement']:
                                ws3 = wb.create_sheet("Par département")
                                df_depts = pd.DataFrame(report['statistiques_par_departement'])
                                for r_idx, row in enumerate(df_depts.itertuples(), start=1):
                                    for c_idx, value in enumerate(row[1:], start=1):
                                        cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                                        if r_idx == 1:
                                            cell.fill = header_fill
                                            cell.font = header_font
                            
                            # Sauvegarder en mémoire
                            excel_buffer = BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            st.download_button(
                                "📊 Télécharger rapport Excel",
                                excel_buffer.getvalue(),
                                f"rapport_recherche_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        except Exception as e:
                            st.warning(f"⚠️ Export Excel non disponible: {e}")
                    else:
                        st.info("💡 Installez openpyxl pour l'export Excel: `pip install openpyxl`")
                
        except Exception as e:
            st.error(f"❌ Erreur génération rapport: {e}")
            import traceback
            st.code(traceback.format_exc())

# ✅ Initialiser les variables GitHub Actions dans session_state AVANT de les utiliser
if 'github_workflow_id' not in st.session_state:
    st.session_state.github_workflow_id = None
if 'github_workflow_status' not in st.session_state:
    st.session_state.github_workflow_status = None
if 'github_workflow_conclusion' not in st.session_state:
    st.session_state.github_workflow_conclusion = None
# Initialiser aussi scraping_running si nécessaire (pour la vérification ci-dessous)
if 'scraping_running' not in st.session_state:
    st.session_state.scraping_running = False

# ✅ CRITIQUE : Vérifier si on a un workflow GitHub Actions actif au démarrage
# Si oui, maintenir scraping_running = True pour garder le dashboard visible
# Note: get_github_workflow_status est défini plus tard, donc on ne peut pas l'appeler ici
# Cette vérification sera faite dans la section du dashboard GitHub Actions

# ✅ Options avancées
with st.expander("⚙️ Options avancées"):
    col_adv1, col_adv2 = st.columns(2)
    with col_adv1:
        # ✅ API activée par défaut avec filtre <100k habitants
        use_api_communes = st.checkbox(
            "Utiliser API data.gouv.fr pour les communes",
            value=True,  # ✅ Activé par défaut
            help="✅ RECOMMANDÉ : Récupère automatiquement toutes les communes depuis l'API officielle avec coordonnées GPS. Si désactivé, utilise le fichier data/villes_par_departement.json (liste limitée)"
        )
        if use_api_communes:
            # ✅ Valeurs par défaut : 0 à 100k habitants
            min_pop = st.number_input("Population minimum", min_value=0, value=0, step=100, help="Villes avec au moins cette population")
            max_pop = st.number_input("Population maximum", min_value=0, value=100000, step=1000, help="Villes avec au maximum cette population (défaut: 100k)")
            
            # ✅ Bouton pour afficher les communes trouvées (uniquement pour l'affichage, pas nécessaire pour le scraping)
            if st.button("📋 Afficher les communes trouvées", help="Affiche la liste des communes qui seront scrapées. Le scraping fonctionne même sans cliquer sur ce bouton."):
                st.session_state.show_communes = True
    with col_adv2:
        enable_resume = st.checkbox(
            "Activer resume/checkpoint",
            value=True,
            disabled=use_github_actions,  # Désactiver pour GitHub Actions
            help="Permet de reprendre le scraping où il s'est arrêté (non disponible avec GitHub Actions)"
        )
        num_threads = st.slider(
            "Nombre de threads",
            min_value=1,
            max_value=20,
            value=10,  # ✅ Défaut à 10
            help="Nombre de navigateurs en parallèle (attention: plus de threads = plus rapide mais plus de ressources)"
        )

# ✅ Afficher les communes si demandé
if st.session_state.get('show_communes', False) and use_api_communes and departements:
    st.markdown("---")
    st.subheader("📍 Communes trouvées via API data.gouv.fr")
    
    communes_trouvees = {}
    with st.spinner("🔄 Récupération des communes depuis l'API..."):
        for dept in departements:
            communes = get_communes_from_api(dept, min_pop if use_api_communes else 0, max_pop if use_api_communes else 50000)
            communes_trouvees[dept] = communes
    
    # Afficher un tableau avec toutes les communes
    all_communes = []
    for dept, communes in communes_trouvees.items():
        for commune in communes:
            all_communes.append({
                'Département': dept,
                'Commune': commune['nom'],
                'Code postal': commune['code_postal'],
                'Population': commune['population'] if commune['population'] > 0 else None
            })
    
    if all_communes:
        st.info(f"📊 Total: {len(all_communes)} communes trouvées")
        
        # ✅ Mise en page côte à côte : tableau et carte
        col_table, col_map = st.columns([1, 1])
        
        with col_table:
            st.subheader("📋 Liste des communes")
            df_communes = pd.DataFrame(all_communes)
            # ✅ Convertir la colonne 'Population' en numérique pour le tri
            if 'Population' in df_communes.columns:
                df_communes['Population'] = pd.to_numeric(df_communes['Population'], errors='coerce').fillna(0).astype(int)
                # Créer une colonne formatée pour l'affichage avec espaces (ex: 56 659)
                # On garde la colonne numérique pour le tri, et on crée une colonne formatée pour l'affichage
                df_communes['Population (formatée)'] = df_communes['Population'].apply(
                    lambda x: f"{x:,}".replace(',', ' ') if x > 0 else "N/A"
                )
                # Pour le tri numérique, on garde la colonne Population comme nombre
                # Pour l'affichage, on utilise la version formatée
                df_communes_display = df_communes[['Département', 'Commune', 'Code postal', 'Population']].copy()
                # On remplace la colonne Population par la version formatée pour l'affichage
                df_communes_display['Population'] = df_communes['Population (formatée)']
            else:
                df_communes_display = df_communes
            
            # CSS pour autofit toutes les colonnes et éviter la colonne vide
            st.markdown("""
            <style>
            /* Cibler le conteneur du DataFrame */
            div[data-testid="stDataFrame"] {
                width: 100% !important;
            }
            div[data-testid="stDataFrame"] > div {
                width: 100% !important;
                overflow-x: visible !important;
            }
            /* Tableau avec ajustement automatique - largeur 100% mais colonnes auto */
            div[data-testid="stDataFrame"] table {
                width: 100% !important;
                table-layout: auto !important;
                border-collapse: collapse !important;
            }
            /* Colonnes avec ajustement automatique selon le contenu */
            div[data-testid="stDataFrame"] th,
            div[data-testid="stDataFrame"] td {
                white-space: nowrap !important;
                padding: 8px 12px !important;
                width: auto !important;
                max-width: none !important;
            }
            /* Masquer complètement la dernière colonne si elle est vide */
            div[data-testid="stDataFrame"] table thead tr th:last-child:empty,
            div[data-testid="stDataFrame"] table tbody tr td:last-child:empty {
                display: none !important;
                width: 0 !important;
                padding: 0 !important;
                border: none !important;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # Utiliser dataframe avec formatage de la population (espaces pour séparer les milliers)
            st.dataframe(df_communes_display, height=600)
            
            # JavaScript pour masquer la colonne vide après le rendu
            st.markdown("""
            <script>
            function hideEmptyColumn() {
                const tables = document.querySelectorAll('div[data-testid="stDataFrame"] table');
                tables.forEach(function(table) {
                    const rows = table.querySelectorAll('tr');
                    if (rows.length > 0) {
                        // Vérifier si la dernière colonne est vide dans toutes les lignes
                        let allEmpty = true;
                        rows.forEach(function(row) {
                            const lastCell = row.querySelector('th:last-child, td:last-child');
                            if (lastCell && lastCell.textContent && lastCell.textContent.trim() !== '') {
                                allEmpty = false;
                            }
                        });
                        
                        // Si toutes les dernières colonnes sont vides, les masquer
                        if (allEmpty) {
                            rows.forEach(function(row) {
                                const lastCell = row.querySelector('th:last-child, td:last-child');
                                if (lastCell) {
                                    lastCell.style.display = 'none';
                                    lastCell.style.width = '0';
                                    lastCell.style.padding = '0';
                                    lastCell.style.border = 'none';
                                }
                            });
                        }
                    }
                });
            }
            // Exécuter après le chargement
            setTimeout(hideEmptyColumn, 100);
            setTimeout(hideEmptyColumn, 500);
            setTimeout(hideEmptyColumn, 1000);
            </script>
            """, unsafe_allow_html=True)
        
        with col_map:
            st.subheader("🗺️ Carte interactive")
            
            # ✅ Carte interactive avec folium
            try:
                import folium
                from streamlit_folium import folium_static
                
                # Filtrer les communes avec coordonnées GPS directement depuis communes_trouvees
                communes_avec_gps = []
                for dept, communes_list in communes_trouvees.items():
                    for comm in communes_list:
                        if comm.get('latitude') and comm.get('longitude'):
                            communes_avec_gps.append({
                                'nom': comm['nom'],
                                'departement': dept,
                                'code_postal': comm.get('code_postal', ''),
                                'population': comm.get('population', 0),
                                'latitude': comm['latitude'],
                                'longitude': comm['longitude']
                            })
                
                if communes_avec_gps:
                    # Calculer le centre de la carte (moyenne des coordonnées)
                    avg_lat = sum(c['latitude'] for c in communes_avec_gps) / len(communes_avec_gps)
                    avg_lon = sum(c['longitude'] for c in communes_avec_gps) / len(communes_avec_gps)
                    
                    # Créer une carte centrée sur les communes avec un meilleur style
                    m = folium.Map(
                        location=[avg_lat, avg_lon], 
                        zoom_start=8,
                        tiles='OpenStreetMap'  # Style plus propre
                    )
                    
                    # ✅ AFFICHER TOUTES LES COMMUNES (pas de limite de 200)
                    populations = [c['population'] for c in communes_avec_gps if c['population'] > 0]
                    
                    if populations:
                        min_pop_displayed = min(populations)
                        max_pop_displayed = max(populations)
                        pop_range_displayed = max_pop_displayed - min_pop_displayed if max_pop_displayed > min_pop_displayed else 1
                    else:
                        min_pop_displayed = 0
                        max_pop_displayed = 1
                        pop_range_displayed = 1
                    
                    # ✅ Améliorer le design : utiliser des clusters pour les grandes quantités
                    from folium.plugins import MarkerCluster
                    marker_cluster = MarkerCluster().add_to(m)
                    
                    for commune in communes_avec_gps:
                        pop = commune['population']
                        pop_str = f"{pop:,}" if pop > 0 else "N/A"
                        popup_text = f"""
                        <div style='font-family: Arial, sans-serif;'>
                            <h4 style='margin: 0 0 10px 0; color: #2c3e50;'>{commune['nom']}</h4>
                            <p style='margin: 5px 0;'><strong>Département:</strong> {commune['departement']}</p>
                            <p style='margin: 5px 0;'><strong>Code postal:</strong> {commune['code_postal']}</p>
                            <p style='margin: 5px 0;'><strong>Population:</strong> {pop_str}</p>
                        </div>
                        """
                        
                        # ✅ Taille du marqueur proportionnelle à la population (plus petite pour un meilleur design)
                        if pop > 0 and pop_range_displayed > 0:
                            # Normaliser entre 3 et 10 pixels de radius
                            normalized = (pop - min_pop_displayed) / pop_range_displayed
                            radius = 3 + (normalized * 7)  # Entre 3 et 10 pixels
                        else:
                            radius = 3
                        
                        # Couleur selon la population (seuils fixes pour la couleur)
                        if pop > 10000:
                            icon_color = '#e74c3c'  # Rouge
                        elif pop > 5000:
                            icon_color = '#f39c12'  # Orange
                        elif pop > 2000:
                            icon_color = '#3498db'  # Bleu
                        else:
                            icon_color = '#27ae60'  # Vert
                        
                        folium.CircleMarker(
                            location=[commune['latitude'], commune['longitude']],
                            radius=radius,
                            popup=folium.Popup(popup_text, max_width=250),
                            tooltip=f"{commune['nom']} ({pop:,} hab.)" if pop > 0 else commune['nom'],
                            color='white',
                            weight=1.5,
                            fillColor=icon_color,
                            fillOpacity=0.7,
                        ).add_to(marker_cluster)
                    
                    # Afficher la carte
                    folium_static(m, width=700, height=600)
                    
                    st.success(f"🗺️ {len(communes_avec_gps)} communes affichées avec coordonnées GPS")
                    
                    # Légende améliorée
                    st.markdown("""
                    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; border-radius: 8px; margin-top: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);'>
                        <strong style='font-size: 16px;'>📊 Légende :</strong><br>
                        <div style='margin-top: 10px; line-height: 1.8;'>
                            <span style='color: #e74c3c; font-weight: bold;'>●</span> Rouge : > 10 000 hab. | 
                            <span style='color: #f39c12; font-weight: bold;'>●</span> Orange : 5 000 - 10 000 hab. | 
                            <span style='color: #3498db; font-weight: bold;'>●</span> Bleu : 2 000 - 5 000 hab. | 
                            <span style='color: #27ae60; font-weight: bold;'>●</span> Vert : < 2 000 hab.<br>
                            <small style='opacity: 0.9;'>La taille des points est proportionnelle à la population. Les points proches sont regroupés automatiquement.</small>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.warning("⚠️ Aucune commune avec coordonnées GPS trouvée")
            except ImportError:
                st.info("💡 Pour afficher une carte interactive, installez: `pip install folium streamlit-folium`")
            except Exception as e:
                st.error(f"Erreur lors de la création de la carte: {e}")
    else:
        st.warning("Aucune commune trouvée avec les critères sélectionnés")
    
    if st.button("❌ Fermer l'affichage des communes"):
        st.session_state.show_communes = False
        st.experimental_rerun()

st.markdown("---")

# État du scraping
if 'scraper' not in st.session_state:
    st.session_state.scraper = None
if 'scraping_running' not in st.session_state:
    st.session_state.scraping_running = False
if 'scraped_results' not in st.session_state:
    st.session_state.scraped_results = []
if 'scraping_thread' not in st.session_state:
    st.session_state.scraping_thread = None
if 'scraping_started' not in st.session_state:
    st.session_state.scraping_started = False
if 'saved_count' not in st.session_state:
    st.session_state.saved_count = 0
if 'logs_buffer' not in st.session_state:
    st.session_state.logs_buffer = []
# Note: github_workflow_id et github_workflow_status sont initialisés plus tôt dans le code

# ✅ CRITIQUE : Vérifier si on a un workflow GitHub Actions actif au démarrage
# Si oui, maintenir scraping_running = True pour garder le dashboard visible
# Cette vérification se fera plus tard, après le chargement de la config GitHub

# ✅ Fonctions pour GitHub Actions
def trigger_github_workflow(token, repo, metiers, departements, max_results, num_threads, use_api_communes, min_pop, max_pop):
    """Déclenche le workflow GitHub Actions"""
    try:
        # ✅ D'abord, récupérer la liste des workflows pour trouver le bon ID
        workflows_url = f"https://api.github.com/repos/{repo}/actions/workflows"
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        
        # Récupérer les workflows
        workflows_response = requests.get(workflows_url, headers=headers)
        if workflows_response.status_code != 200:
            return False, f"Erreur récupération workflows: {workflows_response.status_code} - {workflows_response.text}"
        
        workflows_data = workflows_response.json()
        workflow_id = None
        
        # Chercher le workflow "Google Maps Scraping" ou "scraping.yml"
        for workflow in workflows_data.get('workflows', []):
            if workflow.get('name') == 'Google Maps Scraping' or workflow.get('path', '').endswith('scraping.yml'):
                workflow_id = workflow.get('id')
                break
        
        if not workflow_id:
            # Essayer avec le nom du fichier directement
            url = f"https://api.github.com/repos/{repo}/actions/workflows/scraping.yml/dispatches"
        else:
            # Utiliser l'ID du workflow (plus fiable)
            url = f"https://api.github.com/repos/{repo}/actions/workflows/{workflow_id}/dispatches"
        
        data = {
            "ref": "main",  # Essayer "main" d'abord
            "inputs": {
                "metiers": json.dumps(metiers),
                "departements": json.dumps(departements),
                "max_results": str(max_results),
                "num_threads": str(num_threads),
                "use_api_communes": str(use_api_communes).lower(),
                "min_pop": str(min_pop),
                "max_pop": str(max_pop)
            }
        }
        
        response = requests.post(url, headers=headers, json=data)
        
        # Si 404 avec "main", essayer "master"
        if response.status_code == 404 and data["ref"] == "main":
            data["ref"] = "master"
            response = requests.post(url, headers=headers, json=data)
        
        if response.status_code == 204:
            # ✅ Récupérer le run_id du workflow qui vient d'être lancé
            # Attendre un peu pour que GitHub crée le run
            import time
            time.sleep(2)
            
            # Récupérer le dernier run du workflow
            runs_url = f"https://api.github.com/repos/{repo}/actions/workflows/scraping.yml/runs?per_page=1"
            runs_response = requests.get(runs_url, headers=headers)
            if runs_response.status_code == 200:
                runs_data = runs_response.json().get('workflow_runs', [])
                if runs_data:
                    run_id = runs_data[0].get('id')
                    return True, f"Workflow déclenché avec succès (Run ID: {run_id})", run_id
            
            return True, "Workflow déclenché avec succès"
        else:
            error_msg = response.text
            if response.status_code == 404:
                error_msg += f"\n💡 Vérifiez que le workflow existe dans .github/workflows/scraping.yml et qu'il est commité sur GitHub"
            return False, f"Erreur: {response.status_code} - {error_msg}", None
    except Exception as e:
        return False, f"Erreur: {str(e)}", None

def get_github_workflow_status(token, repo, workflow_id=None):
    """Récupère le statut du workflow GitHub Actions"""
    try:
        if workflow_id:
            url = f"https://api.github.com/repos/{repo}/actions/runs/{workflow_id}"
        else:
            # Récupérer le dernier run (non annulé si possible)
            url = f"https://api.github.com/repos/{repo}/actions/workflows/scraping.yml/runs?per_page=10"
        
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            if workflow_id:
                data = response.json()
            else:
                # Récupérer tous les runs et trouver le dernier non-annulé (ou le dernier tout court)
                runs = response.json().get('workflow_runs', [])
                if not runs:
                    return None, None, None
                
                # Chercher le dernier run non-annulé
                for run in runs:
                    if run.get('conclusion') != 'cancelled':
                        data = run
                        break
                else:
                    # Si tous sont annulés, prendre le dernier
                    data = runs[0]
            
            status = data.get('status')  # queued, in_progress, completed
            conclusion = data.get('conclusion')  # success, failure, cancelled, etc.
            run_id = data.get('id')
            return status, conclusion, run_id
        else:
            return None, None, None
    except Exception as e:
        logger.error(f"Erreur récupération statut: {e}")
        return None, None, None

def get_github_workflow_logs(token, repo, run_id):
    """Récupère les logs du workflow GitHub Actions"""
    try:
        # Récupérer les jobs du workflow
        url = f"https://api.github.com/repos/{repo}/actions/runs/{run_id}/jobs"
        headers = {
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            jobs = response.json().get('jobs', [])
            logs = []
            for job in jobs:
                if job.get('status') == 'completed':
                    # Récupérer les logs du job
                    logs_url = job.get('logs_url')
                    if logs_url:
                        logs_response = requests.get(logs_url, headers=headers)
                        if logs_response.status_code == 200:
                            # Les logs sont dans un format spécifique GitHub
                            logs.append({
                                'job_name': job.get('name'),
                                'status': job.get('conclusion'),
                                'logs_url': logs_url
                            })
            return logs
        return []
    except Exception as e:
        logger.error(f"Erreur récupération logs: {e}")
        return []

# ✅ NOTE : Les fonctions list_github_workflows, cancel_github_workflow et cancel_all_github_workflows 
# sont maintenant définies plus haut (ligne ~206) pour éviter NameError

# ✅ Cette section a été déplacée en haut pour être visible dès le démarrage (voir ligne ~200)

# ✅ Boutons de contrôle SIMPLIFIÉS (comme demandé)
col_btn1, col_btn2 = st.columns(2)

with col_btn1:
    # ✅ GitHub Actions uniquement - pas de mode local
    # ✅ Le bouton LANCER est toujours activé - on peut lancer plusieurs workflows en parallèle
    button_text = "☁️ LANCER"
    
    if st.button(button_text, disabled=False):
        if not metiers or not departements:
            st.error("⚠️ Veuillez sélectionner au moins un métier et un département")
        elif not github_token or not github_repo:
            st.error("⚠️ Veuillez renseigner le token GitHub et le repository")
        else:
            # ✅ Vérifier les villes déjà scrapées
            villes_deja_scrapees = []
            villes_a_scraper = []
            
            # Préparer la liste des villes à scraper
            for dept in departements:
                if use_api_communes:
                    communes = get_communes_from_api(dept, min_pop if use_api_communes else 0, max_pop if use_api_communes else 50000)
                    villes_dept = [c['nom'] for c in communes]
                else:
                    villes_dept = villes_par_dept.get(dept, [])
                
                for metier in metiers:
                    for ville in villes_dept:
                        if is_already_scraped(metier, dept, ville):
                            villes_deja_scrapees.append(f"{metier} - {dept} - {ville}")
                        else:
                            villes_a_scraper.append(f"{metier} - {dept} - {ville}")
            
            # Afficher un avertissement si certaines villes sont déjà scrapées
            if villes_deja_scrapees:
                st.warning(f"⚠️ {len(villes_deja_scrapees)} combinaison(s) métier/département/ville déjà scrapée(s). Elles seront ignorées si vous continuez.")
                with st.expander("📋 Voir les villes déjà scrapées"):
                    for v in villes_deja_scrapees[:20]:  # Limiter à 20 pour l'affichage
                        st.text(v)
                    if len(villes_deja_scrapees) > 20:
                        st.caption(f"... et {len(villes_deja_scrapees) - 20} autres")
            
            if not villes_a_scraper:
                st.error("❌ Toutes les combinaisons sélectionnées ont déjà été scrapées. Veuillez sélectionner d'autres options.")
            else:
                st.info(f"✅ {len(villes_a_scraper)} combinaison(s) à scraper")
                
                # Déclencher le workflow GitHub Actions
                result = trigger_github_workflow(
                    github_token,
                    github_repo,
                    metiers,
                    departements,
                    max_results,
                    num_threads,
                    use_api_communes,
                    min_pop if use_api_communes else 0,
                    max_pop if use_api_communes else 50000
                )
                
                # ✅ Gérer le retour (peut être (success, message) ou (success, message, run_id))
                if len(result) == 3:
                    success, message, run_id = result
                else:
                    success, message = result
                    run_id = None
                
                if success:
                    st.success(f"✅ {message}")
                    st.info("⏳ Le scraping est en cours sur GitHub Actions. Les résultats sont sauvegardés directement dans la BDD.")
                    st.session_state.scraping_running = True
                    st.session_state.github_workflow_status = 'in_progress'
                    st.session_state.departements_selected = departements
                    st.session_state.metiers_selected = metiers
                    
                    # ✅ Stocker le run_id si disponible
                    if run_id:
                        st.session_state.github_workflow_id = run_id
                    # ✅ Marquer qu'on vient de lancer un workflow pour ne pas l'annuler
                    st.session_state.workflow_just_launched = True
                    st.experimental_rerun()
                else:
                    st.error(f"❌ {message}")

with col_btn2:
    # ✅ Bouton ARRÊTER (simplifié - pas de confirmation)
    if github_token and github_repo:
        if st.button("⏹️ ARRÊTER", help="Arrêter tous les workflows GitHub Actions en cours", key="stop_all_workflows"):
            with st.spinner("⏹️ Arrêt des workflows..."):
                success, message = cancel_all_github_workflows(github_token, github_repo)
                if success:
                    st.success(f"✅ {message}")
                    st.session_state.scraping_running = False
                    st.session_state.github_workflow_status = None
                    st.session_state.github_workflow_id = None
                    st.session_state.scraped_results = []
                    st.experimental_rerun()
                else:
                    st.error(f"❌ {message}")

# ✅ NOTE : Le dashboard GitHub Actions a été supprimé car les workflows sont maintenant gérés en haut de la page
# avec le bouton "Rafraîchir les workflows" qui affiche les statistiques pour chaque workflow

# Zone de scraping (ancien dashboard supprimé - tout est géré en haut maintenant)
# Le code du dashboard GitHub Actions a été supprimé car il n'est plus nécessaire
# Tous les workflows sont maintenant gérés en haut avec le bouton "Rafraîchir les workflows"

# Ancien code du dashboard complètement supprimé

# ✅ Vérifier si on doit annuler les workflows au démarrage
# ✅ NE PAS annuler si on vient juste de lancer un workflow
if github_token and github_repo and not st.session_state.get('workflow_just_launched', False):
    # ✅ Tuer automatiquement tous les workflows en cours au démarrage (une seule fois)
    if not st.session_state.get('workflows_cancelled_on_start', False):
        with st.spinner("⏹️ Annulation des workflows GitHub Actions en cours..."):
            success, message = cancel_all_github_workflows(github_token, github_repo)
            # ✅ TOUJOURS définir le flag pour éviter les boucles, même en cas d'échec
            st.session_state.workflows_cancelled_on_start = True
            if success:
                # Ne pas réinitialiser scraping_running si on a un workflow_id
                if not st.session_state.get('github_workflow_id'):
                    st.session_state.scraping_running = False
                    st.session_state.github_workflow_status = None
                    st.session_state.github_workflow_id = None
                st.success(f"✅ {message}")
                # ✅ Rerun seulement si on a annulé avec succès
                st.experimental_rerun()
            else:
                st.warning(f"⚠️ {message}")
                # ✅ NE PAS faire de rerun en cas d'échec pour éviter les boucles
# ✅ Réinitialiser le flag après le premier rerun
if st.session_state.get('workflow_just_launched', False):
    st.session_state.workflow_just_launched = False

# ✅ GitHub Actions uniquement - plus de code local

# ✅ CRITIQUE : Charger automatiquement les résultats depuis le JSON ET la BDD au démarrage
# Si on n'a pas de résultats en session_state, essayer de les charger depuis le JSON et la BDD
if not st.session_state.scraped_results:
    results_file = Path(__file__).parent.parent.parent / "data" / "scraping_results_github_actions.json"
    results_list = []
    
    # 1. Charger depuis le fichier JSON (si existe)
    if results_file.exists():
        try:
            with open(results_file, 'r', encoding='utf-8') as f:
                results_data = json.load(f)
                if isinstance(results_data, dict) and 'results' in results_data:
                    results_list = results_data['results']
                elif isinstance(results_data, list):
                    results_list = results_data
        except Exception as e:
            logger.error(f"Erreur chargement JSON: {e}")
    
    # 2. ✅ AUSSI charger depuis la BDD (pour voir les résultats sauvegardés directement)
    try:
        from whatsapp_database.queries import get_artisans
        artisans_bdd = get_artisans(limit=10000)  # Récupérer tous les artisans
        if artisans_bdd:
            # Convertir les artisans de la BDD en format compatible
            for artisan in artisans_bdd:
                # Éviter les doublons (par téléphone)
                if not any(r.get('telephone') == artisan.get('telephone') for r in results_list if r.get('telephone')):
                    # ✅ Extraire département depuis code_postal si manquant
                    dept = artisan.get('departement')
                    code_postal = artisan.get('code_postal')
                    if not dept and code_postal:
                        code_postal_str = str(code_postal).strip()
                        if len(code_postal_str) >= 2:
                            if code_postal_str.startswith('97') or code_postal_str.startswith('98'):
                                dept = code_postal_str[:3]
                            else:
                                dept = code_postal_str[:2]
                    
                    # ✅ Si toujours pas de département, essayer depuis ville_recherche via API
                    if not dept and artisan.get('ville_recherche'):
                        try:
                            ville_nom = artisan.get('ville_recherche', '').strip()
                            if ville_nom:
                                url = f"https://geo.api.gouv.fr/communes?nom={ville_nom}&fields=codeDepartement&limit=1"
                                response = requests.get(url, timeout=3)
                                if response.status_code == 200:
                                    communes = response.json()
                                    if communes and len(communes) > 0:
                                        dept = communes[0].get('codeDepartement', '')
                        except:
                            pass  # Si l'API échoue, on continue sans département
                    
                    results_list.append({
                        'nom': artisan.get('nom_entreprise') or artisan.get('nom'),
                        'telephone': artisan.get('telephone'),
                        'site_web': artisan.get('site_web'),
                        'google_maps_url': artisan.get('google_maps_url'),
                        'adresse': artisan.get('adresse'),
                        'ville': artisan.get('ville'),
                        'code_postal': code_postal,
                        'departement': dept,
                        'note': artisan.get('note'),
                        'nb_avis': artisan.get('nombre_avis'),
                        'ville_recherche': artisan.get('ville_recherche'),
                        'recherche': artisan.get('type_artisan') or artisan.get('recherche')
                    })
    except Exception as e:
        logger.error(f"Erreur chargement BDD: {e}")
    
    if results_list:
        st.session_state.scraped_results = results_list

# Afficher les résultats scrapés
if st.session_state.scraped_results:
    st.markdown("---")
    st.subheader("📊 Résultats scrapés")
    
    df = pd.DataFrame(st.session_state.scraped_results)
    
    # Stats
    avec_tel = len(df[df['telephone'].notna()])
    avec_site = len(df[df['site_web'].notna()])
    sans_site = len(df[df['site_web'].isna()])
    
    col_res1, col_res2, col_res3, col_res4 = st.columns(4)
    with col_res1:
        st.metric("Total scrapés", len(df))
    with col_res2:
        st.metric("Avec téléphone", f"{avec_tel} ({avec_tel/len(df)*100:.1f}%)")
    with col_res3:
        st.metric("Avec site web", f"{avec_site} ({avec_site/len(df)*100:.1f}%)")
    with col_res4:
        st.metric("⭐ SANS site web", f"{sans_site} ({sans_site/len(df)*100:.1f}%)")
    
    # ✅ Filtres supprimés comme demandé
    df_filtre = df.copy()
    
    # ✅ Afficher le tableau avec toutes les colonnes importantes
    colonnes_afficher = ['nom', 'telephone', 'site_web', 'google_maps_url', 'adresse', 'ville', 'code_postal', 'departement', 'ville_recherche', 'note', 'nb_avis']
    colonnes_disponibles = [col for col in colonnes_afficher if col in df_filtre.columns]
    
    # CSS amélioré pour le tableau
    css_table = (
        "<style>"
        ".stDataFrame { width: 100% !important; }"
        ".stDataFrame > div { width: 100% !important; }"
        ".stDataFrame table { width: 100% !important; table-layout: auto !important; }"
        ".stDataFrame th, .stDataFrame td { padding: 8px !important; word-wrap: break-word !important; overflow-wrap: break-word !important; white-space: normal !important; }"
        ".stDataFrame th:nth-child(1) { width: 15% !important; }"
        ".stDataFrame th:nth-child(2) { width: 10% !important; }"
        ".stDataFrame th:nth-child(3) { width: 20% !important; }"
        ".stDataFrame th:nth-child(4) { width: 20% !important; }"
        ".stDataFrame th:nth-child(5) { width: 12% !important; }"
        ".stDataFrame th:nth-child(6) { width: 10% !important; }"
        ".stDataFrame th:nth-child(7) { width: 6% !important; }"
        ".stDataFrame th:nth-child(8) { width: 7% !important; }"
        "</style>"
    )
    st.markdown(css_table, unsafe_allow_html=True)
    
    # ✅ Utiliser width au lieu de use_container_width (compatibilité Streamlit)
    st.dataframe(
        df_filtre[colonnes_disponibles],
        height=600
    )
    
    # ✅ Bouton d'export CSV complet
    csv_all = df.to_csv(index=False, encoding='utf-8-sig')
    dept = st.session_state.get('departements_selected', departements)[0] if st.session_state.get('departements_selected') else (departements[0] if departements else '77')
    metier_export = st.session_state.get('metiers_selected', metiers)[0] if st.session_state.get('metiers_selected') else (metiers[0] if metiers else 'plombier')
    st.download_button(
        "📥 Télécharger CSV complet",
        csv_all,
        f"{metier_export}_{dept}_complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        "text/csv"
    )
    
# ✅ Expander "Mode d'exécution" tout en bas de la page
with st.expander("ℹ️ Mode d'exécution", expanded=False):
    st.info("☁️ **Mode GitHub Actions activé** - Le scraping s'exécutera sur GitHub Actions (gratuit jusqu'à 2000 min/mois)")
    st.info("ℹ️ Le scraping s'exécutera sur GitHub Actions. Les résultats sont sauvegardés directement dans la BDD en temps réel.")
